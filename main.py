import threading
from customtkinter import *
from openpyxl.drawing.image import Image
from PIL import Image
from openpyxl.styles import Alignment
import os
import re
import pyglet
import ttkbootstrap as tb
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import openpyxl.utils as utils
import time
import ttkbootstrap as ttk
from tkinter import font
from openpyxl.utils import range_boundaries

app = CTk()
app.geometry("480x580+630+75")
app.title("Excel Tool")
app.configure(fg_color="#FFFFFF")
app.resizable(False, False) 
app.title("Excel Tool")

def move_window(event): # Moving the window
    app.geometry(f'+{event.x_root}+{event.y_root}')

var_fileName = StringVar()
varName = StringVar()
progress_var = DoubleVar()
steps = 2000
uploads = []
filePathsAll = []
signatures = []
clicks = 0

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


titleLableImage = CTkImage(light_image=Image.open("imgs/heading.png"), dark_image=Image.open("imgs/heading.png"), size=(212, 45))

titleLable = CTkLabel(app,text="",image=titleLableImage ) 
titleLable.place(relx=0.5, rely=0.1, anchor="center")

bi = CTkImage(light_image=Image.open("imgs/frame.png"), dark_image=Image.open("imgs/frame.png"), size=(425, 208))

mylable = CTkLabel(app, text="", image=bi) 

mylable.place(relx=0.5, rely=0.35, anchor="center")
mylable.lift()
frame = CTkScrollableFrame(master=app, width=210, border_color="#000000",fg_color="#FFFFFF", 
                           scrollbar_fg_color="#FFFFFF", 
                           scrollbar_button_color="#FFFFFF", scrollbar_button_hover_color="#FFFFFF")

border = CTkButton(master=app, text="", height=3, 
                width=420, text_color="#1E1E1E",
                fg_color='#D9D9D9',
                hover_color="#D9D9D9",
                
                )
underline = Font(name='Calibri', size=18, bold=True, underline='double')
app.columnconfigure(0, weight=1)
app.rowconfigure(0, weight=1)
app.rowconfigure(1, weight=1)
frame.grid(row=1, column=0, sticky='swe')




frame.columnconfigure(0, weight=1)
frame.columnconfigure(1, weight=10)



space_col = CTkLabel(frame,text="", height=20)
space_col.grid(row=0, column=1,  columnspan=3, sticky='n')

space_row = CTkLabel(frame,text="", width=20)
space_row.grid(row=0, column=0,  rowspan=9, sticky='n')

excelLogo = CTkImage(light_image=Image.open("imgs/excellogo.png"), dark_image=Image.open("imgs/excellogo.png"), size=(20, 20))


def hasIntroductionSheet(pathToExcelFile:str )->bool:
    try:
        workbook = openpyxl.load_workbook(pathToExcelFile)
        sheetNames = workbook.sheetnames 
        if 'Introduction' in sheetNames:return 1
        else:return 0

    except KeyError as e:
        OpenErrorWindow(e)
        
    except FileNotFoundError as e:
        OpenErrorWindow(e)


def getSignaturesFromAFMSheet(pathToExcelFile:str):
    try:
        global signatures
        workbook = openpyxl.load_workbook(pathToExcelFile)
        afmSheet = workbook['Sheet1']
        for i, rows in enumerate(afmSheet.iter_rows()): 
            for cell_value in rows:
                if str(cell_value.value):
                    font = cell_value.font
                    if ("_" in str(cell_value.value)) or ("-" in str(cell_value.value)):
                        signatures.append(cell_value.value)


        workbook.save(pathToExcelFile)
        return 
    except KeyError as e:
        OpenErrorWindow(e)
        
    except FileNotFoundError as e:
        OpenErrorWindow(e)


def setAFMStatus(pathToExcelFile:str):
    try:
        workbook = openpyxl.load_workbook(pathToExcelFile)
        summarySheet = workbook['Summary']
        messageNames = []
        AfmStatus = {}
        names = {}
        indices = {}
        namesInSignatures = {}
        cnt = 0
        # print(summarySheet.merged_cells.ranges)
        for l in range(1, 100):
            cell = summarySheet[f'K{l}']
            font = cell.font
            if cell.value=='SKY NETWORK SERVICES Action':
                cnt+=1
                summarySheet[f'K{l}'].value = ''
                summarySheet[f'K{l}'].value = 'AFM \nStatus'
                summarySheet[f'K{l}'].fill = PatternFill(start_color= 'E5E4E2', end_color="E5E4E2", fill_type="solid")
                summarySheet[f'K{l}'].font = Font(bold=True)
                summarySheet[f'K{l}'].alignment = Alignment(horizontal='center', vertical='center')
                summarySheet[f'K{l}'].border = thin_border

                summarySheet[f'L{l}'].value = 'SKY\nNETWORK\nSERVICES\nAction'
                summarySheet[f'L{l}'].fill = PatternFill(start_color= 'E5E4E2', end_color="E5E4E2", fill_type="solid")
                summarySheet[f'L{l}'].font = Font(bold=True)
                summarySheet[f'K{l}'].alignment = Alignment(horizontal='center', vertical='center')
                summarySheet[f'L{l}'].border = thin_border

            
                    
        for j, rows in enumerate(summarySheet.iter_rows()): 
            if (rows[0].hyperlink) :
                startRow = rows[0].row
                # print(startRow)
                messageNames.append(rows[0].value)
                if rows[0].value in names:
                    names[rows[0].value]+=1
                else:
                    names[rows[0].value] = 1

                if rows[0].value in indices:
                    indices[rows[0].value].append(startRow)
                else:
                    indices[rows[0].value] = [startRow]

        for i in messageNames:
            if i in signatures:
                AfmStatus[i] = 'Yes'
            else:
                AfmStatus[i] = 'No'

        for i in messageNames:
            for j in indices[i]:
                summarySheet[f'K{j}'].value = AfmStatus[i]
                summarySheet[f'K{j}'].alignment = Alignment(horizontal='center', vertical='center')
                summarySheet[f'K{j}'].border = thin_border
                summarySheet[f'L{j}'].border = thin_border
                
        workbook.save(pathToExcelFile)
        return
    except KeyError as e:
        OpenErrorWindow(e)
        
        
    except FileNotFoundError as e:
        OpenErrorWindow(e)
        print(f"Error loading workbook {pathToExcelFile}: {e}")




def removeSheet(pathToExcelFile:str, sheetName:str):
    try:
        workbook = openpyxl.load_workbook(pathToExcelFile)
        sheetNames = workbook.sheetnames
        if sheetName in sheetNames:
            sheetToRemove = workbook[sheetName]
            workbook.remove(sheetToRemove)
        workbook.save(pathToExcelFile)
        return

        
    except KeyError as e:
        OpenErrorWindow(e)
        
    except FileNotFoundError as e:
        OpenErrorWindow(e)


def editIntroductionSheet(pathToExcelFile:str, filename:str):
    workbook = openpyxl.load_workbook(pathToExcelFile)  
    try:
        if not hasIntroductionSheet(pathToExcelFile):
            workbook.create_sheet('Introduction', 0)
            sheetNames = workbook.sheetnames

            ti_c = os.path.getctime(pathToExcelFile)
            modified_date = time.ctime(ti_c)
            date_obj = datetime.strptime(modified_date, '%a %b %d %H:%M:%S %Y')
            output_date_str = date_obj.strftime('%d %B %y')
            
            IntroductionSheet = workbook['Introduction']
            Name = os.path.basename(pathToExcelFile)
            fileName = os.path.splitext(Name)[0]
            
            for row in IntroductionSheet.iter_rows():
                for cell in row:
                    cell.border = None
            img = openpyxl.drawing.image.Image('imgs/Picture1.jpg')
            img.width =  (14*96)//2.24
            img.height = (10*96)//2.24
            
            IntroductionSheet.add_image(img, 'B2')
            header_font = Font(bold=True, color= '000000' , size=16, underline='single')
            header_fill = PatternFill(start_color= 'E5E4E2', end_color="E5E4E2", fill_type="solid")
            headFill = PatternFill(start_color= 'DCDCDC', end_color="DCDCDC", fill_type="solid")
            cellFill = PatternFill(start_color= 'F5F5F5', end_color="F5F5F5", fill_type="solid")
            bold = Font(bold=True)
            IntroductionSheet["C26"].value = filename
            IntroductionSheet["C26"].font = header_font
            IntroductionSheet.row_dimensions[26].height = 40

            IntroductionSheet["C26"].alignment = Alignment(horizontal='center', vertical='center')
            
            IntroductionSheet['B28'].value = 'Prepared for:'
            IntroductionSheet['B28'].fill = headFill
            IntroductionSheet['B28'].font = bold
            IntroductionSheet['C28'].value = 'Customer Name:'
            IntroductionSheet['C28'].fill = cellFill
            IntroductionSheet['D28'].value = 'SKY NETWORKS'
            
            IntroductionSheet['C29'].value = 'Customer Contact '
            IntroductionSheet['C29'].fill = cellFill
            IntroductionSheet['D29'].value = ' '
            IntroductionSheet['C30'].value = 'Customer Email'
            IntroductionSheet['C30'].fill = cellFill
            IntroductionSheet['D30'].value = ' '
            IntroductionSheet['C31'].value = 'Customer Phone'
            IntroductionSheet['C31'].fill = cellFill
            IntroductionSheet['D31'].value = ' '
            
            IntroductionSheet['B34'].value = 'Prepared by:'
            IntroductionSheet['B34'].fill = headFill
            IntroductionSheet['B34'].font = bold
            IntroductionSheet['C34'].value = 'Cisco Contact'
            IntroductionSheet['C34'].fill = cellFill
            IntroductionSheet['D34'].value = 'Balaji Dhanasekar'
            IntroductionSheet['C35'].value = 'Cisco Email'
            IntroductionSheet['C35'].fill = cellFill
            IntroductionSheet['D35'].value = 'badhanas@cisco.com'
            IntroductionSheet['C36'].value = 'Cisco Phone'
            IntroductionSheet['C36'].fill = cellFill
            IntroductionSheet['D36'].value = ' +91 9600111374'
            IntroductionSheet['B38'].value = 'DCP Reference:'
            IntroductionSheet['B38'].fill = headFill
            IntroductionSheet['B38'].font = bold
            IntroductionSheet['C38'].value = ' '
            IntroductionSheet['B39'].value = 'Classification'
            IntroductionSheet['B39'].fill = cellFill
            IntroductionSheet['C39'].value = 'Cisco Highly Confidential'
            
            IntroductionSheet['B40'].value = 'PID:'
            IntroductionSheet['B40'].fill = cellFill
            IntroductionSheet['C40'].value = ' '
            
            IntroductionSheet['B41'].value = 'Template Version:'
            IntroductionSheet['B41'].fill = cellFill
            IntroductionSheet['C41'].value = 'v1.0'
            IntroductionSheet['A43'].value = ' Document History:'
            IntroductionSheet['A43'].fill = headFill
            IntroductionSheet['A43'].font = bold
            IntroductionSheet['B43'].value = 'Date:'
            IntroductionSheet['B43'].fill = cellFill
            IntroductionSheet['B43'].font = bold
            IntroductionSheet['B44'].value = output_date_str

            IntroductionSheet['C43'].value = 'Author:'
            IntroductionSheet['C43'].fill = cellFill
            IntroductionSheet['C43'].font = bold
            IntroductionSheet['C44'].value = 'Balaji Dhanasekar'
            IntroductionSheet['D43'].value = 'Version:'
            IntroductionSheet['D43'].fill = cellFill
            IntroductionSheet['D43'].font = bold
            IntroductionSheet['D44'].value = '1.0'
            IntroductionSheet['E43'].value = 'Comments'
            IntroductionSheet['E43'].fill = cellFill
            IntroductionSheet['E43'].font = bold
            IntroductionSheet['E44'].value = 'Final Draft'

            for row in IntroductionSheet.iter_rows(max_col=50, max_row=100):
                
                for cell in row:
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                    if cell.value is None:
                        cell.border = Border(left=Side(style='thin', color='FFFFFF'), 
                        right=Side(style='thin', color='FFFFFF'), 
                        top=Side(style='thin', color='FFFFFF'), 
                        bottom=Side(style='thin', color='FFFFFF'))
                    else:
                        if cell.value == fileName:
                            continue
                        cell.border = Border(left=Side(style='thin', color='000000'), 
                        right=Side(style='thin',color='000000'), 
                        top=Side(style='thin',color='000000'), 
                        bottom=Side(style='thin',color='000000'))
                        cell.fill=cellFill

            for col in ['B','C','D']:
                IntroductionSheet[f'{col}26'].fill = header_fill
            for col in ['A', 'B', 'C','D','E', 'F', 'G']:
                IntroductionSheet.column_dimensions[f'{col}'].width = 30
                
            workbook.save(pathToExcelFile)
            sheetNames = workbook.sheetnames
  
        else:
            return
    
    except KeyError as e:
        OpenErrorWindow(e)
        
        workbook.save(pathToExcelFile)
    return

def close():
    app.destroy()

def editExcelFile(pathToExcelFile):
    try:
        Name = os.path.basename(pathToExcelFile)
        fileName = os.path.splitext(Name)[0]
        workbook = openpyxl.load_workbook(pathToExcelFile)
        varName.set('Uploading')
        btn.configure(state="disabled")
        global clicks
        clicks+=1
        frame.rowconfigure(clicks, weight=1)
        # var_fileName.set(fileName)
        SelectedfileName = CTkLabel(frame,text_color="#45484D",fg_color='#ffffff',  text=fileName, height=32, width=260, anchor="w", font=(r'Sf-pro\SFPRODISPLAYREGULAR.OTF', 12))
        logolable = CTkLabel(frame, text='', image=excelLogo)#, 
        logolable.grid(row=str(clicks), column='0', padx='25', sticky='nswe')
        SelectedfileName.grid(row=str(clicks), column='1', padx='0', sticky='we')
        SelectedfileName.lift()
        

        if fileName:
            submitButton.configure(state='normal')
            
            
        if 'Italy' in fileName:
            uploads.append('Italy')
            # toast.show_toast()
            removeSheet(pathToExcelFile, 'OtherMessages')
            removeSheet(pathToExcelFile, 'NonActionableMessages')
            editIntroductionSheet(pathToExcelFile, fileName)
            if 'AFM Signatures' in uploads:
                setAFMStatus(pathToExcelFile)
                varName.set('Upload')
                btn.configure(state="normal")
                # pass
            else:
                varName.set('Upload')
                btn.configure(state="normal")
                return
            
        if 'AFM Signatures' in fileName:
            uploads.append('AFM Signatures')
            # toast.show_toast()
            getSignaturesFromAFMSheet(pathToExcelFile)
            varName.set('Upload')
            btn.configure(state="normal")

            return

            
        if 'Telco' in fileName:
            uploads.append('Telco')
            # toast.show_toast()
            removeSheet(pathToExcelFile, 'OtherMessages')
            removeSheet(pathToExcelFile, 'NonActionableMessages')
            editIntroductionSheet(pathToExcelFile, fileName)

            if 'AFM Signatures' in uploads:
                setAFMStatus(pathToExcelFile)
                varName.set('Upload')

                btn.configure(state="normal")

                # pass
            else:
                varName.set('Upload')
                btn.configure(state='normal')
                return

        
        if 'No Lab' in fileName:
            uploads.append('Telco')
            # toast.show_toast()
            removeSheet(pathToExcelFile, 'OtherMessages')
            removeSheet(pathToExcelFile, 'NonActionableMessages')
            editIntroductionSheet(pathToExcelFile, fileName)
            if 'AFM Signatures' in uploads:
                setAFMStatus(pathToExcelFile)
                varName.set('Upload')
                btn.configure(state="normal")

                # pass
            else:
                varName.set('Upload')
                btn.configure(state='normal')
                return   

    except KeyError as e:
        OpenErrorWindow(e)
        # print(f"Error accessing worksheet: {e}", "editExcelFile")
        

    except FileNotFoundError as e:
        OpenErrorWindow(f"Error loading workbook {pathToExcelFile}: {e}")
        print(f"Error loading workbook {pathToExcelFile}: {e}")

def startLoadingThread():
    loadingThread = threading.Thread(target=open_window)
    loadingThread.start()

def OpenErrorWindow(error):
    error_window=CTkToplevel(app)
    error_window.title("Error")
    error_window.geometry("350x150+680+350") 
    error_window.grab_set()
    error_window.focus()
    error_window.pack_slaves()
    error_window.update()
    erroeMsg = CTkLabel(error_window, text=error)
    erroeMsg.place(relx=0.5, rely=0.5,anchor="center")
    return


def open_window():
    top_level=CTkToplevel(app)
    top_level.title("Processing")
    top_level.geometry("350x150+680+350") 
    progress_bar = CTkProgressBar(top_level,orientation="horizontal", 
                                  width=250, 
                                  height=15,
                                  corner_radius=0,
                                  progress_color="#4BB543",
                                  fg_color="#F0F0F0",
                                  variable=progress_var)
    progress_bar.place(relx=0.5, rely=0.43, anchor="center")
   
    progress_bar.set(0)
    top_level.grab_set()
    top_level.focus()
    top_level.pack_slaves()

    top_level.update()
    steps = 50
    progressVal = 1/(steps)
    stepVal=0
    for i in range(steps):
        
        top_level.update()
        stepVal= stepVal+progressVal
        time.sleep(0.05)
        progress_bar.set(stepVal)

    if int(stepVal)==1:
        top_level.after(4000, top_level.destroy())

       
def selectFile():
    global filePathsAll
    filePath = filedialog.askopenfile(filetypes=[("Excel Files", "*.xlsx;*.xls")])      
    if filePath:
        
        entry = CTkEntry(master=app)
        entry.insert(0, filePath)  
        dir =os.path.abspath(filePath.name)   
        afmURL = ''
        # print(filePath.name)
        filePathsAll.append(filePath.name)
        editExcelFile(filePath.name)
        varName.set("Select File")

        btn.configure(state="normal")
        print(filePathsAll)

varName.set("Select Files")
btn = CTkButton(master=app, text="Select files", height=37, 
                width=110, text_color="#000000",
                border_color="#A9ACB4",
                fg_color='#FFFFFF',font=(r'Sf-pro\SFPRODISPLAYREGULAR.OTF', 12),
                hover_color="#F9F9F9",
                border_width=1.2,
                corner_radius=7, textvariable=varName,command=selectFile)   

submitButton =  CTkButton(master=app, text="Start editing", height=37, 
                width=110, text_color="#FFFFFF",
                border_color="#A9ACB4",
                fg_color='#262626',
                hover_color="#1e1e1e",
                border_width=1.2,
                corner_radius=7, 
                state='disabled',
                command=open_window)  

cancelButton = CTkButton(master=app, text="Close", height=37, 
                width=110, text_color="#1E1E1E",
                border_color="#A9ACB4",
                fg_color='#FFFFFF',
                hover_color="#F9F9F9",
                border_width=1.2,
                corner_radius=7, command=close)  

submitButton.place(relx=0.84, rely=0.94, anchor="center")
cancelButton.place(relx=0.18, rely=0.94, anchor="center")

btn.place(relx=0.5, rely=0.46, anchor="center")

border.place(relx=0.5, rely=0.88 ,anchor='center')

btn.lift()
border.lift()

    
app.mainloop()